from flask import Flask, request, jsonify
from flask_cors import CORS
import requests
import json
import pandas as pd
import openpyxl as xl
import networkx as nx
import matplotlib.pyplot as plt

app = Flask(__name__)
CORS(app)

# Replace with your API base URL
API_BASE_URL = "https://twitter-api45.p.rapidapi.com/"

# Replace with your API keys or tokens if required
HEADERS = {
    "x-rapidapi-host": "twitter-api45.p.rapidapi.com",
    "x-rapidapi-key": "a9f5d0c888msha36716e833304e7p1bdbb3jsncd693dddd494"
}

json_obj = {}

@app.route('/network', methods=['POST'])
def network_analysis():
    try:
        # Receive JSON data (assuming the format is {'tweet_id': '12345'})
        data = request.get_json()
        print(f"{data}")
        tweet_id = data.get("tweet_id")
        print(f"Processing tweet_id: {tweet_id}")

        # Fetch retweeters for the tweet ID
        retweeters = get_top_retweeters(tweet_id)

        # Prepare the response by extracting screen names
        retweeter_data = [{"screen_name": retweeter["screen_name"]} for retweeter in retweeters["retweets"] if "screen_name" in retweeter]

        global json_obj
        json_obj = {
            "retweeters": retweeter_data  # Include only screen names in the response
        }
        print(f"Final response: {json_obj}")

        # Convert JSON to Excel and create graph
        excel_file = 'output.xlsx'
        convert_json_to_excel(json_obj, excel_file)
        create_graph(excel_file)

        return jsonify(json_obj)

    except Exception as e:
        print(f"Error in network_analysis: {str(e)}")
        return jsonify({"error": str(e)}), 500


def get_top_retweeters(tweet_id):
    try:
        url = f"{API_BASE_URL}retweets.php?id={tweet_id}"
        response = requests.get(url, headers=HEADERS)

        if response.status_code == 200:
            retweeters = response.json()  # Parse the JSON
            return retweeters  # Return the full list of retweeters
        else:
            print(f"API error {response.status_code} for tweet {tweet_id}: {response.text}")
            return []  # Return an empty list on error
    except requests.exceptions.RequestException as e:
        print(f"Request failed for retweeters: {e}")
        return []  # Return an empty list on exception


def convert_json_to_excel(json_obj, excel_file):
    """Converts JSON data to an Excel file."""
    retweeters = json_obj['retweeters']
    
    # Create a DataFrame from the retweeter list
    df_retweeters = pd.DataFrame(retweeters)  # Directly use the list of dictionaries
    df_retweeters.to_excel(excel_file, sheet_name="Sheet1", index=False)

    print(f"Data successfully converted to {excel_file}")


def create_graph(excel_file):
    """Creates a graph from an Excel file linking the main culprit to users."""
    
    print(f"Loading Excel file: {excel_file}")  # Debug print

    G = nx.DiGraph()  # Use a directed graph for arrows
    
    # Define the main culprit
    main_culprit = "main"

    # Load the Excel file and read the first sheet (Retweets)
    workbook = xl.load_workbook(excel_file)
    sheet1 = workbook[workbook.sheetnames[0]]
    
    # Extract user data from the first sheet
    users = []  # List to hold user data
    for row in sheet1.iter_rows(values_only=True, min_row=2):
        user = row[0]  # User from the first column, assuming it has the screen names
        if user:  # Ensure user is not None or empty
            users.append(user)

    # Add the main culprit node in the center
    G.add_node(main_culprit)

    # Link each user to the main culprit
    for user in users:
        G.add_node(user)  # Add user node
        G.add_edge(main_culprit, user)  # Link user to main culprit

    # Plot the graph with colors
    pos = nx.spring_layout(G)  # Automatic layout for better spacing
    plt.figure(figsize=(10, 10))
    
    # Define node colors and edge colors
    node_colors = ['lightgreen' if node == main_culprit else 'lightblue' for node in G.nodes()]
    edge_colors = ['orange' for _ in G.edges()]

    # Draw the graph
    nx.draw(G, pos, with_labels=True, node_size=2000, node_color=node_colors, font_size=10, font_color="black", font_weight="bold", arrows=True)
    nx.draw_networkx_edges(G, pos, edge_color=edge_colors, arrowstyle='-|>', arrowsize=20)

    plt.title("Graph of Main Banda and Users")
    plt.show()


if __name__ == '__main__':
    app.run(debug=True, port=5100)

